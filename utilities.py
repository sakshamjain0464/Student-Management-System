'''Module for processing data from frontend and sending it to backed'''

from tkinter import *                                  #for GUI
from CTkMessagebox import CTkMessagebox as message     #for displaying prompt messages
import backend                            #for database access  'Self Created'
import re                 #for validating entries
from openpyxl import Workbook, chart         #for accessing and creating excel sheets
import os                                  #for accessing folders
from pathlib import Path                   #for getting path of folders


# Function to create user
def createUser(create_window, user_entry, pwd_entry, conf_pwd_entry):
    '''Funcion to process data for creating user  and sending datato backend '''
    if(user_entry.get() == "" or pwd_entry.get() == "" or conf_pwd_entry.get() == ""):
        message(title="WARNING", message="Kindly Fill all the fields!", icon='warning')

    elif(pwd_entry.get() != conf_pwd_entry.get()):
        message(title="WARNING", message="Password and Confirm Password Does not Match!", icon='warning')
    
    else:
        if(message(title="Confirm", message="Do you want to proceed!", option_2 = 'YES', option_1 = 'NO', icon='question').get() == 'YES'):
            created = backend.createUserToDb(user_entry.get(), pwd_entry.get())
            if(created == True):
                message(title="Created", message="User Created Successfully!", icon='check')
            elif(created == "exists"):
                message(title="Already Exists", message="User Already Exists!", icon='info')
            else:
                message(title="WARNING", message="Username or password must be unique!", icon='warning')
            user_entry.delete(0, END)
            pwd_entry.delete(0, END)
            conf_pwd_entry.delete(0, END)
        else:
            pass


# Function to add new student
def addStudent(first_name, last_name, course, cgpa):
    '''Function to process and validate data for creating student and sending datato backend'''

    if(first_name.get() == "" or last_name.get() == "" or course.get() == "" or cgpa.get() == ''):
        message(title="WARNING", message="Kindly Fill all the fields!", icon='warning')

    elif(bool(re.search("\s", first_name.get())) or bool(re.search("\s", last_name.get())) or bool(re.search("\s", cgpa.get()))):
        message(title="WARNING", message="Any field cannot contain spaces", icon='warning')

    elif(any(not c.isalpha() for c in first_name.get()) or any(not c.isalpha() for c in last_name.get())):
        message(title="WARNING", message="First Name or Last Name cannot contain numbers and special characters", icon='warning')

    elif(not cgpa.get().isnumeric()):
        message(title="WARNING", message="CGPA can only be a two digit integer ", icon='warning')

    elif(int(cgpa.get()) < 0 or int(cgpa.get()) > 100):
        message(title="WARNING", message="CGPA can not be less than 0 or greater than 10", icon='warning')

    else:
        if(message(title="Confirm", message="Do you want to proceed!", option_2 = 'YES', option_1 = 'NO', icon='question').get() == 'YES'):
            new_roll = f"S{str(backend.getLastRollNumber()+1)}"
            added = backend.addStudentToDb(new_roll, first_name.get(), last_name.get(), course.get(), int(cgpa.get()))
            if(added == True):
                message(title="Created", message=f"Student Added Successfully! for Roll No {new_roll}", icon='check')
                first_name.delete(0, END)
                last_name.delete(0, END)
                cgpa.delete(0, END)
            elif(added == 'Error'):
                message(title="Error", message="An Error Occurred", icon='cancel')
            else:
                message(title="Error", message="Cannot add Student, contact the administrator", icon='cancel')
        else:
            pass

def getCourses():
    '''Geting Course codes from db and sending to frontend returns a list of course codes'''
    return backend.getCoursesFromDb()

def fetchStudentDetails(roll):
    '''Fetching student details from db and sending to frontend returns a list of details'''
    if(roll == ""):
        message(title="WARNING", message="Kindly Enter Roll Number!", icon='warning')
        return False
    
    elif(roll[0] != 'S' or any(not c.isalnum() for c in roll)):
        message(title="WARNING", message="Kindly Enter a valid Roll Number!", icon='warning')
        return False
    
    else:
        details = backend.fetchStudentDetailsFromDb(roll)
        if(details == False):
             message(title="WARNING", message="Student details not found!", icon='cancel')
             return False
        elif(details == "error"):
            message(title="Error", message="An Error Occurred", icon='cancel')
            return False
        else:
            return details
        

def updateName(details, first_name, last_name):
    '''Function to process and validate data for updating student name and sending datato backend'''
    if(first_name.get() == "" or last_name.get() == ""):
        message(title="WARNING", message="Kindly Fill all the fields!", icon='warning')

    elif(bool(re.search("\s", first_name.get())) or bool(re.search("\s", last_name.get()))):
        message(title="WARNING", message="Any field cannot contain spaces", icon='warning')

    elif(any(not c.isalpha() for c in first_name.get()) or any(not c.isalpha() for c in last_name.get())):
        message(title="WARNING", message="Any field cannot special characters or numbers", icon='warning')

    else:
        if(message(title="Confirm", message="Do you want to proceed!", option_2 = 'YES', option_1 = 'NO', icon='question').get() == 'YES'):
            updated = backend.updateNameToDb(details[0], first_name.get(), last_name.get())
            if(updated == True):
                message(title="Updated", message=f"Student Name Successfully Updated!", icon='check')
                first_name.delete(0, END)
                last_name.delete(0, END)
            elif(updated == 'Error'):
                message(title="Error", message="An Error Occurred", icon='cancel')
            else:
                message(title="Error", message="Cannot update Student, contact the administrator", icon='cancel')
        else:
            pass
        

def updateCourse(details, course):
    '''Function to process and validate data for updating student course and sending datato backend'''
    if(message(title="Confirm", message="Do you want to proceed!", option_2 = 'YES', option_1 = 'NO', icon='question').get() == 'YES'):
        updated = backend.updateCourseToDb(details[0], course.get())
        if(updated == True):
            message(title="Updated", message=f"Course Successfully Updated!", icon='check')
        elif(updated == 'Error'):
             message(title="Error", message="An Error Occurred", icon='cancel')
        else:
            message(title="Error", message="Cannot update Student, contact the administrator", icon='cancel')

    else:
        pass



def updateCGPA(details, cgpa):
    '''Function to process and validate data for updating student name and sending datato backend'''
    if(cgpa.get() == ''):
        message(title="WARNING", message="Kindly Fill all the fields!", icon='warning')

    elif(not cgpa.get().isnumeric()):
        message(title="WARNING", message="CGPA can only be a two digit integer ", icon='warning')

    elif(int(cgpa.get()) < 0 or int(cgpa.get()) > 100):
        message(title="WARNING", message="CGPA can not be less than 0 or greater than 10", icon='warning')

    else:
        if(message(title="Confirm", message="Do you want to proceed!", option_2 = 'YES', option_1 = 'NO', icon='question').get() == 'YES'):
            updated = backend.updateCGPAToDb(details[0], int(cgpa.get()))
            if(updated == True):
                message(title="Updated", message=f"Course Successfully Updated!", icon='check')
            elif(updated == 'Error'):
                 message(title="Error", message="An Error Occurred", icon='cancel')
            else:
                message(title="Error", message="Cannot update Student, contact the administrator", icon='cancel')
        else:
            pass



def deleteStudent(details):
    ''''Function to send delete request to backend'''

    if(message(title="Confirm", message="Do you want to proceed!", option_2 = 'YES', option_1 = 'NO', icon='question').get() == 'YES'):
        deleted = backend.deleteFromDb(details[0])
        if(deleted == True):
            message(title="Updated", message=f"Student Deleted Successfully!", icon='check')
        elif(deleted == 'Error'):
            message(title="Error", message="An Error Occurred", icon='cancel')
        else:
            message(title="Error", message="Cannot update Student, contact the administrator", icon='cancel')

    else:
        pass



def addCourse(id, name):
    '''Function to process and validate data for adding course name and sending datato backend'''
    if(id.get() == "" or name.get() == ''):
        message(title="WARNING", message="Kindly Fill all the fields!", icon='warning')

    elif(bool(re.search("\s", id.get()))):
        message(title="WARNING", message="Course ID field cannot contain spaces", icon='warning')

    elif(any(not c.isalnum() for c in id.get())):
        message(title="WARNING", message="Course ID cannot contain special characters", icon='warning')

    elif(not id.get().isupper()):
        message(title="WARNING", message="Course ID should be upper case ", icon='warning')

    elif(len(id.get()) > 5):
        message(title="WARNING", message="Course ID can only have 5 characters", icon='warning')

    else:
        if(message(title="Confirm", message="Do you want to proceed!", option_2 = 'YES', option_1 = 'NO', icon='question').get() == 'YES'):
            added = backend.addCourseToDb(id.get(), name.get())
            if(added == True):
                message(title="Created", message=f"Course Added SuccessFully", icon='check')
            elif(added == 'Error'):
                message(title="Error", message="An Error Occurred", icon='cancel')
            else:
                message(title="Error", message="Cannot add Student, contact the administrator", icon='cancel')
        else:
            pass



def fetchCourseDetails(courseid):
    '''Function for fetching course details from backend'''
    if(courseid == ""):
        message(title="WARNING", message="Kindly Enter Course ID!", icon='warning')
        return False
    else:
        details = backend.fetchCourseDetailsFromDb(courseid)
        if(details == False):
             message(title="WARNING", message="Course details not found!", icon='cancel')
             return False
        elif(details == "error"):
            message(title="Error", message="An Error Occurred", icon='cancel')
            return False
        else:
            return details
        
def updateCourseName(details, name):
    '''Function to process and validate data for updating course name and sending datato backend'''
    if(name.get() == ""):
        message(title="WARNING", message="Kindly Fill all the fields!", icon='warning')

    else:
        if(message(title="Confirm", message="Do you want to proceed!", option_2 = 'YES', option_1 = 'NO', icon='question').get() == 'YES'):
            updated = backend.updateCourseNameToDb(details[0], name.get())
            if(updated == True):
                message(title="Updated", message=f"Course Name Successfully Updated!", icon='check')
                name.delete(0, END)
                name.delete(0, END)
            elif(updated == 'error'):
                message(title="Error", message="An Error Occurred", icon='cancel')
            else:
                message(title="Error", message="Cannot update Student, contact the administrator", icon='cancel')
        else:
            pass

def getCourseReport():
    '''Functiion for fetching data from backend to generate course report excel shee tand save it to downloads folder'''
    report =backend.getCourseReportFromDb()

    if(report == "error"):
        message(title="Error", message="An Error Occurred", icon='cancel')

    else:
        #initilaizing sheet
        wb = Workbook()                 #created blank workbook
        sheet = wb.active               #create blank sheet

        #setting up sheet
        sheet.column_dimensions['B'].width = 50
        sheet.column_dimensions['A'].width = 20

        #adding headers to sheet
        sheet.cell(row=1, column=1).value = "Course Code"
        sheet.cell(row=1, column=2).value = "Course Name"
        sheet.cell(row=1, column=3).value = "No. of Students"

        # putting data in sheet
        for data in report:
            print(data)
            sheet.append(data)

        # Initializing report chart
        mychart = chart.BarChart()
        mychart.type = "col"
        mychart.title = "Performance of Course"
        mychart.y_axis.title = 'No of Students'
        mychart.x_axis.title = 'Course'
        mychart.legend = None

        students = chart.Reference(sheet, min_col=3, min_row=2, max_row=sheet.max_row, max_col=3)
        courses = chart.Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row, max_col=1)

        # Putting data on chart
        mychart.add_data(students)
        mychart.set_categories(courses)
        
        # Add chart on sheet
        sheet.add_chart(mychart, "E2")

        # saving data to sheet
        try:
            downloads = os.path.join(str(Path.home()), 'Downloads')            #getting downloads folder path
            wb.save(f"{downloads}\\CourseReport.xlsx")
            message(title="Created", message=f"Report Generated SuccessFully in Downloads Folder", icon='check')
        except:
            message(title="Error", message="Cannot Generate Report!", icon='cancel')

def close():
    '''Function for closing application'''
    backend.connection.close()
    message(title="Error", message="Thank You for using our System!", icon='check')

